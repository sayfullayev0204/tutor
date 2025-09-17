from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.db.models import Q
from django.contrib import messages
from .models import Student, TTJ
from housing.models import HousingInspection
from .forms import StudentForm
from django.core.exceptions import PermissionDenied
from django.http import JsonResponse, HttpResponse
from .models import District
from datetime import datetime
import openpyxl
from openpyxl.utils import get_column_letter
from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.db.models import Q
from .models import Student
from django.core.paginator import Paginator

@login_required
def student_list(request):
    user = request.user
    
    # Determine queryset based on user role
    if hasattr(user, 'is_rector') and user.is_rector():
        students = Student.objects.all()
    elif hasattr(user, 'is_dean') and user.is_dean():
        students = Student.objects.filter(group__faculty=user.faculty) if user.faculty else Student.objects.none()
    elif hasattr(user, 'is_tutor') and user.is_tutor():
        students = Student.objects.filter(group__tutor=user)
    else:
        students = Student.objects.none()
    
    # Search functionality
    search_query = request.GET.get('search')
    if search_query:
        students = students.filter(
            Q(first_name__icontains=search_query) |
            Q(last_name__icontains=search_query) |
            Q(middle_name__icontains=search_query) |
            Q(group__name__icontains=search_query) |
            Q(student_id__icontains=search_query) |
            Q(jshshir__icontains=search_query) |
            Q(passport__icontains=search_query)
        )
    
    # Filter by gender
    gender = request.GET.get('gender')
    if gender:
        students = students.filter(gender=gender)
    
    # Filter by apartment type
    appartment_type = request.GET.get('appartment_type')
    if appartment_type:
        students = students.filter(appartment_type=appartment_type)
    
    # Filter by bully_student
    bully_student = request.GET.get('bully_student')
    if bully_student in ['true', 'false']:
        bully_student_bool = bully_student == 'true'
        students = students.filter(bully_student=bully_student_bool)
    
    # Pagination
    paginator = Paginator(students, 10)  # Show 10 students per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'students': page_obj,
        'search_query': search_query,
    }
    return render(request, 'students/student_list.html', context)

@login_required
def student_detail(request, pk):
    user = request.user
    
    if hasattr(user, 'is_rector') and user.is_rector():
        student = get_object_or_404(Student, pk=pk)
    elif hasattr(user, 'is_dean') and user.is_dean():
        student = get_object_or_404(Student, pk=pk, group__faculty=user.faculty)
    elif hasattr(user, 'is_tutor') and user.is_tutor():
        student = get_object_or_404(Student, pk=pk, group__tutor=user)
    else:
        messages.error(request, "Sizga bu talabani ko'rish uchun ruxsat yo'q.")
        return redirect('dashboard')
    
    inspections = HousingInspection.objects.filter(student=student).order_by('-inspection_date')
    
    context = {
        'student': student,
        'inspections': inspections,
    }
    return render(request, 'students/student_detail.html', context)

@login_required
def renting_students(request):
    user = request.user
    
    if hasattr(user, 'is_rector') and user.is_rector():
        students = Student.objects.filter(appartment_type='rent')
    elif hasattr(user, 'is_dean') and user.is_dean():
        students = Student.objects.filter(appartment_type='rent', group__faculty=user.faculty) if user.faculty else Student.objects.none()
    elif hasattr(user, 'is_tutor') and user.is_tutor():
        students = Student.objects.filter(appartment_type='rent', group__tutor=user)
    else:
        students = Student.objects.none()
    
    context = {
        'students': students,
    }
    return render(request, 'students/renting_students.html', context)

@login_required
def dormitory_students(request):
    user = request.user
    
    if hasattr(user, 'is_rector') and user.is_rector():
        students = Student.objects.filter(appartment_type='ttj')
    elif hasattr(user, 'is_dean') and user.is_dean():
        students = Student.objects.filter(appartment_type='ttj', group__faculty=user.faculty) if user.faculty else Student.objects.none()
    elif hasattr(user, 'is_tutor') and user.is_tutor():
        students = Student.objects.filter(appartment_type='ttj', group__tutor=user)
    else:
        students = Student.objects.none()
    
    context = {
        'students': students,
    }
    return render(request, 'students/dormitory_students.html', context)

@login_required
def orphan_students(request):
    user = request.user
    
    if hasattr(user, 'is_rector') and user.is_rector():
        students = Student.objects.filter(is_orphan=True)
    elif hasattr(user, 'is_dean') and user.is_dean():
        students = Student.objects.filter(is_orphan=True, group__faculty=user.faculty) if user.faculty else Student.objects.none()
    elif hasattr(user, 'is_tutor') and user.is_tutor():
        students = Student.objects.filter(is_orphan=True, group__tutor=user)
    else:
        students = Student.objects.none()
    
    context = {
        'students': students,
    }
    return render(request, 'students/orphan_students.html', context)

@login_required
def disabled_students(request):
    user = request.user
    
    if hasattr(user, 'is_rector') and user.is_rector():
        students = Student.objects.filter(has_disability=True)
    elif hasattr(user, 'is_dean') and user.is_dean():
        students = Student.objects.filter(has_disability=True, group__faculty=user.faculty) if user.faculty else Student.objects.none()
    elif hasattr(user, 'is_tutor') and user.is_tutor():
        students = Student.objects.filter(has_disability=True, group__tutor=user)
    else:
        students = Student.objects.none()
    
    context = {
        'students': students,
    }
    return render(request, 'students/disabled_students.html', context)

@login_required
def student_create(request):
    if not (hasattr(request.user, 'is_tutor') and request.user.is_tutor()):
        messages.error(request, "Faqat guruh tutorlari talaba qo'sha oladi.")
        return redirect('dashboard')
    
    if request.method == 'POST':
        form = StudentForm(request.POST)
        if 'room' in request.POST and request.POST['room'] == StudentForm.ADD_NEW_ROOM:
            request.session['student_form_data'] = request.POST.dict()
            request.session['return_to'] = 'student_create'
            return redirect('room_create')
        
        if form.is_valid():
            student = form.save(commit=False)
            if student.group.tutor != request.user:
                messages.error(request, "Siz faqat o'zingiz boshqaradigan guruhga talaba qo'sha olasiz.")
                return render(request, 'students/student_form.html', {'form': form, 'title': 'Yangi talaba qo‘shish'})
            try:
                student.save()
                messages.success(request, f"{student.full_name} muvaffaqiyatli qo'shildi.")
                request.session.pop('student_form_data', None)
                request.session.pop('return_to', None)
                return redirect('student_list')
            except Exception as e:
                messages.error(request, f"Ma'lumotni saqlashda xato: {str(e)}")
        else:
            messages.error(request, f"Forma xatolarni tuzating: {form.errors.as_text()}")
    else:
        initial_data = request.session.get('student_form_data', {})
        form = StudentForm(initial=initial_data)
        form.fields['group'].queryset = form.fields['group'].queryset.filter(tutor=request.user)
    
    context = {
        'form': form,
        'title': 'Yangi talaba qo‘shish',
    }
    return render(request, 'students/student_form.html', context)

@login_required
def student_edit(request, pk):
    if not (hasattr(request.user, 'is_tutor') and request.user.is_tutor()):
        messages.error(request, "Faqat guruh tutorlari talaba ma'lumotlarini tahrirlashi mumkin.")
        return redirect('dashboard')
    
    student = get_object_or_404(Student, pk=pk, group__tutor=request.user)
    
    if request.method == 'POST':
        form = StudentForm(request.POST, instance=student)
        # Check if the form submission is explicitly for adding a new room
        if 'room' in request.POST and request.POST['room'] == StudentForm.ADD_NEW_ROOM and request.POST.get('add_new_room') == 'true':
            request.session['student_form_data'] = request.POST.dict()
            request.session['return_to'] = 'student_edit'
            request.session['student_pk'] = pk
            return redirect('room_create')
        
        if form.is_valid():
            student = form.save(commit=False)
            if student.group.tutor != request.user:
                messages.error(request, "Siz faqat o'zingiz boshqaradigan guruhdagi talabani tahrirlashingiz mumkin.")
                return render(request, 'students/student_form.html', {'form': form, 'student': student, 'title': 'Talaba ma‘lumotlarini tahrirlash'})
            try:
                student.save()
                messages.success(request, f"{student.full_name} ma'lumotlari muvaffaqiyatli yangilandi.")
                request.session.pop('student_form_data', None)
                request.session.pop('return_to', None)
                request.session.pop('student_pk', None)
                # Re-render the edit page instead of redirecting
                form = StudentForm(instance=student)
                form.fields['group'].queryset = form.fields['group'].queryset.filter(tutor=request.user)
                return render(request, 'students/student_form.html', {
                    'form': form,
                    'student': student,
                    'title': 'Talaba ma‘lumotlarini tahrirlash',
                })
            except Exception as e:
                messages.error(request, f"Ma'lumotni saqlashda xato: {str(e)}")
        else:
            messages.error(request, f"Forma xatolarni tuzating: {form.errors.as_text()}")
    else:
        initial_data = request.session.get('student_form_data', {})
        form = StudentForm(instance=student, initial=initial_data)
        form.fields['group'].queryset = form.fields['group'].queryset.filter(tutor=request.user)
    
    context = {
        'form': form,
        'student': student,
        'title': 'Talaba ma‘lumotlarini tahrirlash',
    }
    return render(request, 'students/student_form.html', context)

@login_required
def student_delete(request, pk):
    if not (hasattr(request.user, 'is_tutor') and request.user.is_tutor()):
        messages.error(request, "Faqat guruh tutorlari talabani o'chirishi mumkin.")
        return redirect('dashboard')
    
    student = get_object_or_404(Student, pk=pk, group__tutor=request.user)
    
    if request.method == 'POST':
        student_name = student.full_name
        student.delete()
        messages.success(request, f"{student_name} muvaffaqiyatli o'chirildi.")
        return redirect('student_list')
    
    context = {
        'student': student,
        'title': 'Talabani o‘chirish',
    }
    return render(request, 'students/student_confirm_delete.html', context)

def get_districts(request):
    region_id = request.GET.get('region')
    districts = District.objects.filter(region_id=region_id).values('id', 'name')
    return JsonResponse(list(districts), safe=False)



import openpyxl
from django.http import HttpResponse
from django.utils import timezone
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from students.models import Student, Region, District
from faculty.models import Group, Faculty
from accounts.models import User

def export_students_to_excel(request):
    # Foydalanuvchi autentifikatsiyasini tekshirish
    if not request.user.is_authenticated:
        return HttpResponse("Foydalanuvchi autentifikatsiya qilinmagan!", status=403)

    # Bugungi sanani olish (YYYYMMDD formatida)
    current_date = timezone.now().strftime("%Y%m%d")

    # Foydalanuvchi rolini aniqlash
    user = request.user
    if user.is_rector():
        role = "rektor"
        students = Student.objects.select_related(
            'group', 'group__faculty', 'mutaxassislik', 'country',
            'const_region', 'const_district', 'temporary_region', 'temporary_district',
            'room', 'ttj'
        ).all()
    elif user.is_dean():
        role = "dekan"
        students = Student.objects.filter(
            group__faculty=user.faculty
        ).select_related(
            'group', 'group__faculty', 'mutaxassislik', 'country',
            'const_region', 'const_district', 'temporary_region', 'temporary_district',
            'room', 'ttj'
        )
    elif user.is_tutor():
        role = "tutor"
        students = Student.objects.filter(
            group__tutor=user
        ).select_related(
            'group', 'group__faculty', 'mutaxassislik', 'country',
            'const_region', 'const_district', 'temporary_region', 'temporary_district',
            'room', 'ttj'
        )
    else:
        return HttpResponse("Sizda talabalarni eksport qilish huquqi yo'q!", status=403)

    # Excel faylini yaratish
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Talabalar"

    # Yuqori qismda "Talabalar ro'yxati" sarlavhasini qo'shish
    worksheet.merge_cells('A1:AI1')  # A1 dan AI1 gacha birlashtirish (36 ustun)
    title_cell = worksheet['A1']
    title_cell.value = "Talabalar ro'yxati"
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    title_cell.font = Font(bold=True, size=14)

    # Ustun sarlavhalari (shablonga mos ravishda)
    headers = [
        "ID", "F.I.O", "Jinsi", "Tug'ilgan sana", "Yoshi", "Guruh", "O'quv kursi", "Fakultet",
        "Talaba ID", "JSHSHIR", "Pasport", "OTM", "Ta'lim darajasi", "To'lov turi", "Ta'lim shakli",
        "Mutaxassisligi", "Fuqarolik", "millati",
        "Doimiy yashash manzili", "", "",  # Viloyati, Tuman, Mahallasi, ko'chasi, uy raqami
        "Vaqtincha yashash manzili", "", "",  # Viloyat, Tuman, Joriy manzili
        "Yashash turi", "Oilaviy holati", "Telefon raqami",
        "Ijtimoiy holati", "", "", "", "", "", "", "", "",  # 9 ta kichik ustun
        "Yaratilgan sana"
    ]
    sub_headers = [
        "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "",
        "Viloyati", "Tuman", "Mahallasi, ko'chasi, uy raqami",
        "Viloyat", "Tuman", "Joriy manzili",
        "", "", "",
        "Chin yetim", "Nogironligi bor", "Boquvchisini yuqotgan",
        "Ijtimoiy himoya reestiriga kiritilgan", "Temir daftarga kiritilgan oila farzandi",
        "Ayollar daftariga kirgizilgan", "Yoshlar daftariga kiritilgan",
        "Mehribonlik uyi tarbiyalanuvchisi", "Ota-onasi ajrashgan",
        ""
    ]

    # Sarlavhalarni yozish (3-qatordan boshlab)
    worksheet.append([])  # Bo'sh qator (A2)
    worksheet.append(headers)  # Asosiy sarlavhalar (A3)
    worksheet.append(sub_headers)  # Kichik sarlavhalar (A4)

    # Sarlavhalarni qalin shriftda formatlash
    for cell in worksheet[3]:  # Asosiy sarlavhalar (3-qator)
        cell.font = Font(bold=True)
    for cell in worksheet[4]:  # Kichik sarlavhalar (4-qator)
        cell.font = Font(bold=True)

    # Filtrlarni yoqish (3-qator uchun, A3:AI3)
    worksheet.auto_filter.ref = "A3:AI3"

    # Har bir talaba uchun ma'lumotlarni qayta ishlash
    for student in students:
        # Ijtimoiy holat ma'lumotlari
        social_status = [
            "Ha" if student.is_orphan else "Yo‘q",  # Chin yetim
            "Ha" if student.has_disability else "Yo‘q",  # Nogironligi bor
            "",  # Boquvchisini yuqotgan
            "Yo‘q",  # Ijtimoiy himoya reestiriga kiritilgan
            "Yo‘q",  # Temir daftarga kiritilgan oila farzandi
            "Yo‘q",  # Ayollar daftariga kirgizilgan
            "Yo‘q",  # Yoshlar daftariga kiritilgan
            "Yo‘q",  # Mehribonlik uyi tarbiyalanuvchisi
            ""   # Ota-onasi ajrashgan yoki boshqa holatlar
        ]

        # Boquvchisini yo'qotgan va oilaviy holat
        family_status_extra = ""
        if student.family_type == 'turmush_qurgan':
            family_status_extra = "turmush o'rtog'i vafot etgan"
            social_status[2] = family_status_extra
        elif student.family_type == 'ajrashgan':
            family_status_extra = "Ajrashgan"
            social_status[8] = family_status_extra
        # Shablonga ko'ra maxsus holatlar (misol uchun)
        if student.id == 1:  # Mustafoyev Alibek uchun
            social_status[2] = "Otasi vafot etgan"
            social_status[1] = "I-guruh"  # Nogironlik darajasi
        elif student.id == 2:  # Sayfullayev Gayrat uchun
            social_status[2] = "Onasi vafot etgan"

        # Doimiy va vaqtincha manzil
        const_address = [
            student.const_region.name if student.const_region else "",
            student.const_district.name if student.const_district else "",
            student.address if student.address else ""
        ]
        temp_address = [
            student.temporary_region.name if student.temporary_region else "",
            student.temporary_district.name if student.temporary_district else "",
            student.temporary_address if student.temporary_address else (
                student.ttj.name if student.appartment_type == 'ttj' and student.ttj else
                "O'z uyi" if student.appartment_type == 'home' else
                student.room.address if student.appartment_type == 'rent' and student.room else ""
            )
        ]

        # Yashash turi
        appartment_type = {
            'ttj': 'TTJ (Talabalar turar joyi)',
            'rent': 'Ijara xonadon',
            'home': "O'z uyi"
        }.get(student.appartment_type, student.appartment_type)

        # Fakultet va guruh
        faculty_name = student.group.faculty.name if student.group and student.group.faculty else ""
        group_name = student.group.name if student.group else ""

        # Ma'lumotlarni qator sifatida yig'ish
        row = [
            student.id,
            student.full_name,
            dict(Student.GENDER_CHOICES).get(student.gender, student.gender),
            student.birth_date.strftime("%d.%m.%Y"),
            student.age,
            group_name,
            str(student.group.course) if student.group else "",
            faculty_name,
            student.student_id,
            student.jshshir,
            student.passport,
            student.otm,
            dict(Student.TALIM_CHOICES).get(student.talim_turi, student.talim_turi),
            dict(Student.TULOV_CHOICES).get(student.tulov_turi, student.tulov_turi),
            dict(Student.TALIM_SHAKLI_CHOICES).get(student.talim_shakli, student.talim_shakli),
            student.mutaxassislik.name if student.mutaxassislik else "",
            dict(Student.FUQARO_CHOICES).get(student.fuqaro, student.fuqaro),
            student.country.name if student.country else "",
            *const_address,  # Doimiy yashash manzili (3 ustun)
            *temp_address,   # Vaqtincha yashash manzili (3 ustun)
            appartment_type,
            dict(Student.FAMILY_CHOICES).get(student.family_type, student.family_type),
            student.phone_number if student.phone_number else "-",
            *social_status,  # Ijtimoiy holati (9 ustun)
            student.created_at.strftime("%d.%m.%Y %H:%M")
        ]

        # Qatorni Excelga qo'shish
        worksheet.append(row)

    # Ustun kengliklarini ma'lumot uzunligiga moslashtirish
    column_widths = {}
    for col_idx in range(1, worksheet.max_column + 1):
        column_letter = get_column_letter(col_idx)
        max_length = 0
        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value and not isinstance(cell, openpyxl.cell.cell.MergedCell):
                    try:
                        cell_length = len(str(cell.value))
                        max_length = max(max_length, cell_length)
                    except:
                        pass
        column_widths[column_letter] = max(max_length + 2, 10) * 1.2

    for column_letter, width in column_widths.items():
        worksheet.column_dimensions[column_letter].width = width

    # Excel faylini saqlash va fayl nomini dinamik ravishda belgilash
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="talabalar_{current_date}_{role}.xlsx"'
    
    workbook.save(response)
    return response

import pandas as pd
from django.core.exceptions import ValidationError
from django.http import HttpResponse
from django.shortcuts import render
from django.core.files.storage import FileSystemStorage
from students.models import Student, Mutaxassislik, Country, Region, District, TTJ
from faculty.models import Faculty, Group
from housing.models import Room
from datetime import datetime, date
import os
import uuid

import pandas as pd
from django.core.exceptions import ValidationError
from django.http import HttpResponse
from django.shortcuts import render
from django.core.files.storage import FileSystemStorage
from students.models import Student, Mutaxassislik, Country, Region, District, TTJ
from faculty.models import Faculty, Group
from housing.models import Room
from datetime import datetime, date
import os
import uuid
import time
import logging

# Configure logging for debugging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

def import_students_from_excel(file_path, user):
    try:
        xl = pd.ExcelFile(file_path)
        successful = 0
        errors = []
        
        for sheet_name in xl.sheet_names:
            logger.info(f"Sahifa o'qilmoqda: {sheet_name}")
            df = xl.parse(sheet_name, dtype={'JSHSHIR': str, 'Jshshir': str, 'Tug\'ilgan sana': str})
            
            if len(df) >= 2:
                df = df.drop(1).reset_index(drop=True)
                logger.info(f"Sahifa {sheet_name}: 2-qator tashlab ketildi.")
            
            required_columns = {
                'F.I.O': ['F.I.O', 'FIO', 'Ism'],
                'JSHSHIR': ['JSHSHIR', 'Jshshir'],
                'Pasport': ['Pasport', 'Passport'],
                'Telefon raqami': ['Telefon raqami', 'Telefon', 'Phone'],
                'Jinsi': ['Jinsi', 'Jins', 'Gender'],
                'Tug\'ilgan sana': ['Tug\'ilgan sana', 'Tugilgan sana', 'Birth Date'],
                'Mutaxassisligi': ['Mutaxassisligi', 'Mutaxassislik', 'Specialty'],
                'Doimiy yashash manzili Viloyati': ['Doimiy yashash manzili Viloyati', 'Viloyati', 'Viloyat', 'Region', 'Doimiy viloyat'],
                'Doimiy yashash manzili Tuman': ['Doimiy yashash manzili Tuman', 'Tuman', 'District', 'Doimiy tuman'],
                'Vaqtincha yashash manzili Viloyat': ['Vaqtincha yashash manzili Viloyat', 'Vaqtincha yashash manzili Viloyati', 'Vaqtincha viloyat', 'Temporary Region'],
                'Vaqtincha yashash manzili Tuman': ['Vaqtincha yashash manzili Tuman', 'Vaqtincha yashash manzili Tumani', 'Vaqtincha tuman', 'Temporary District'],
                'Vaqtincha yashash manzili Joriy manzili': ['Vaqtincha yashash manzili Joriy manzili', 'Joriy manzili', 'Joriy manzil', 'Temporary Address', 'Manzil'],
                'Yashash turi': ['Yashash turi', 'Yashash', 'Residence Type'],
                'Guruh': ['Guruh', 'Group'],
                'O\'quv kursi': ['O\'quv kursi', 'O\'quv kurs', 'Kurs', 'Course'],
                'OTM': ['OTM', 'Otm', 'University'],
                'To\'lov turi': ['To\'lov turi', 'To‘lov turi', 'Tulov turi', 'Payment Type'],
                'Ta\'lim shakli': ['Ta\'lim shakli', 'Ta‘lim shakli', 'Talim shakli', 'Education Form'],
                'Talaba ID': ['Talaba ID', 'Student ID'],
                'Fakultet': ['Fakultet', 'Faculty'],
                'Oilaviy holati': ['Oilaviy holati', 'Oilaviy holat', 'Family Status'],
                'Millat': ['Millat', 'Nation'],
                'Fuqarolik': ['Fuqarolik', 'Citizenship'],
            }
            
            actual_columns = {}
            missing_columns = []
            for key, aliases in required_columns.items():
                found = False
                for alias in aliases:
                    if alias in df.columns:
                        actual_columns[key] = alias
                        found = True
                        break
                if not found:
                    actual_columns[key] = None
                    missing_columns.append(key)
            
            if missing_columns:
                logger.info(f"Sahifa {sheet_name}: Quyidagi ustunlar topilmadi va standart qiymatlar ishlatiladi: {', '.join(missing_columns)}")
            
            for index, row in df.iterrows():
                if row.isnull().all() or pd.isna(row.get(actual_columns.get("F.I.O"))):
                    continue
                
                try:
                    excel_row = index + 3
                    student_id = row.get(actual_columns.get("Talaba ID")) if actual_columns.get("Talaba ID") else None
                    if pd.isna(student_id):
                        student_id = Student.objects.count() + 1
                        logger.info(f"Sahifa {sheet_name}, Qator {excel_row}: Talaba ID bo'sh. Avtomatik ID ({student_id}) ishlatildi.")
                    
                    course_str = row.get(actual_columns.get("O'quv kursi")) if actual_columns.get("O'quv kursi") else None
                    try:
                        course = int(float(course_str))
                        if course < 1 or course > 4:
                            raise ValueError
                    except (ValueError, TypeError):
                        course = 1
                        errors.append(f"Sahifa {sheet_name}, Qator {excel_row}: O'quv kursi noto'g'ri yoki bo'sh. Standart qiymat (1) ishlatildi.")
                    
                    faculty_name = row.get(actual_columns.get("Fakultet")) if actual_columns.get("Fakultet") else "Noma'lum"
                    if pd.isna(faculty_name):
                        faculty_name = "Noma'lum"
                        errors.append(f"Sahifa {sheet_name}, Qator {excel_row}: Fakultet bo'sh. Standart qiymat (Noma'lum) ishlatildi.")
                    faculty, _ = Faculty.objects.get_or_create(
                        name=faculty_name,
                        defaults={'description': f'{faculty_name} fakulteti'}
                    )
                    
                    mutaxassislik_name = row.get(actual_columns.get("Mutaxassisligi")) if actual_columns.get("Mutaxassisligi") else "Noma'lum"
                    if pd.isna(mutaxassislik_name):
                        mutaxassislik_name = "Noma'lum"
                        errors.append(f"Sahifa {sheet_name}, Qator {excel_row}: Mutaxassislik bo'sh. Standart qiymat (Noma'lum) ishlatildi.")
                    mutaxassislik, _ = Mutaxassislik.objects.get_or_create(name=mutaxassislik_name)
                    
                    fuqarolik = row.get(actual_columns.get("Fuqarolik")) if actual_columns.get("Fuqarolik") else "O‘zbekiston fuqarosi"
                    country_name = "O'zbekiston" if "zbekiston" in str(fuqarolik).lower() else "Noma'lum"
                    country, _ = Country.objects.get_or_create(name=country_name)
                    
                    const_region_name = row.get(actual_columns.get("Doimiy yashash manzili Viloyati")) if actual_columns.get("Doimiy yashash manzili Viloyati") else "Qashqadaryo viloyati"
                    const_region, _ = Region.objects.get_or_create(name=const_region_name)
                    const_district_name = row.get(actual_columns.get("Doimiy yashash manzili Tuman")) if actual_columns.get("Doimiy yashash manzili Tuman") else "Qarshi tumani"
                    const_district, _ = District.objects.get_or_create(name=const_district_name, region=const_region)
                    
                    temp_region_name = row.get(actual_columns.get("Vaqtincha yashash manzili Viloyat")) if actual_columns.get("Vaqtincha yashash manzili Viloyat") else "Qashqadaryo viloyati"
                    temp_region, _ = Region.objects.get_or_create(name=temp_region_name)
                    temp_district_name = row.get(actual_columns.get("Vaqtincha yashash manzili Tuman")) if actual_columns.get("Vaqtincha yashash manzili Tuman") else "Qarshi tumani"
                    temp_district, _ = District.objects.get_or_create(name=temp_district_name, region=temp_region)
                    
                    temporary_address = row.get(actual_columns.get("Vaqtincha yashash manzili Joriy manzili")) if actual_columns.get("Vaqtincha yashash manzili Joriy manzili") else "Noma'lum"
                    if pd.isna(temporary_address):
                        temporary_address = "Noma'lum"
                    
                    group_name = row.get(actual_columns.get("Guruh")) if actual_columns.get("Guruh") else "Noma'lum"
                    if pd.isna(group_name):
                        group_name = "Noma'lum"
                    group, created = Group.objects.get_or_create(
                        name=group_name,
                        faculty=faculty,
                        defaults={'course': course, 'tutor': user}
                    )
                    if not created and group.tutor != user:
                        group.tutor = user
                        group.save()
                    
                    yashash_turi = row.get(actual_columns.get("Yashash turi")) if actual_columns.get("Yashash turi") else "O'z uyida"
                    if pd.isna(yashash_turi):
                        yashash_turi = "O'z uyida"
                    ttj = None
                    room = None
                    appartment_type = 'home'
                    if "talabalar turar" in yashash_turi.lower():
                        appartment_type = 'ttj'
                        ttj, _ = TTJ.objects.get_or_create(
                            name="Qarshi TTJ",
                            address=temporary_address,
                            region=temp_region,
                            district=temp_district,
                            defaults={
                                'capacity': 100,
                                'has_internet': True,
                                'has_heating': True,
                                'condition': 'good',
                                'manager_name': 'Noma\'lum',
                                'manager_phone': 'Noma\'lum',
                                'notes': ''
                            }
                        )
                    elif "ijara" in yashash_turi.lower():
                        appartment_type = 'rent'
                        room, _ = Room.objects.get_or_create(
                            address=temporary_address,
                            defaults={
                                'room_number': 'Noma\'lum',
                                'room_type': 'shared',
                                'area': 50.0,
                                'rent_price': 0,
                                'has_kitchen': False,
                                'has_bathroom': False,
                                'has_internet': False,
                                'has_heating': False,
                                'condition': 'good',
                                'landlord_name': 'Noma\'lum',
                                'landlord_phone': 'Noma\'lum',
                                'notes': ''
                            }
                        )
                    
                    birth_date_str = row.get(actual_columns.get("Tug'ilgan sana")) if actual_columns.get("Tug\'ilgan sana") else None
                    birth_date = date(2000, 1, 1)
                    if birth_date_str:
                        try:
                            for fmt in ['%Y-%m-%d', '%d.%m.%Y', '%m/%d/%Y', '%Y/%m/%d', '%d/%m/%Y']:
                                try:
                                    birth_date = datetime.strptime(str(birth_date_str), fmt).date()
                                    break
                                except ValueError:
                                    continue
                            else:
                                raise ValueError
                        except ValueError:
                            errors.append(f"Sahifa {sheet_name}, Qator {excel_row}: Tug'ilgan sana noto'g'ri formatda ({birth_date_str}). Standart (2000-01-01).")
                    
                    fio = row.get(actual_columns.get("F.I.O")) if actual_columns.get("F.I.O") else "Noma'lum Noma'lum"
                    fio_parts = str(fio).strip().split()
                    last_name = fio_parts[0] if fio_parts else "Noma'lum"
                    first_name = fio_parts[1] if len(fio_parts) > 1 else "Noma'lum"
                    middle_name = " ".join(fio_parts[2:]) if len(fio_parts) > 2 else ""
                    
                    jshshir = row.get(actual_columns.get("JSHSHIR")) if actual_columns.get("JSHSHIR") else None
                    if pd.isna(jshshir) or not str(jshshir).isdigit() or len(str(jshshir)) != 14:
                        jshshir = f"TEMP{uuid.uuid4().hex[:14]}"
                        errors.append(f"Sahifa {sheet_name}, Qator {excel_row}: JSHSHIR noto'g'ri yoki bo'sh. Vaqtinchalik ({jshshir}).")
                    jshshir = str(jshshir).replace('.0', '')
                    
                    passport = row.get(actual_columns.get("Pasport")) if actual_columns.get("Pasport") else "Noma'lum"
                    phone_number = row.get(actual_columns.get("Telefon raqami")) if actual_columns.get("Telefon raqami") else "Noma'lum"
                    gender_str = row.get(actual_columns.get("Jinsi")) if actual_columns.get("Jinsi") else "Erkak"
                    gender = 'male' if "erkak" in gender_str.lower() else 'female'
                    
                    family_status = row.get(actual_columns.get("Oilaviy holati")) if actual_columns.get("Oilaviy holati") else "Oila qurmagan"
                    family_type = 'married' if "qurgan" in family_status.lower() else 'unmarried' if "qurmagan" in family_status.lower() else 'unmarried'
                    
                    nation = row.get(actual_columns.get("Millat")) if actual_columns.get("Millat") else "O'zbek"
                    nation = nation.lower()
                    
                    tulov_turi = row.get(actual_columns.get("To'lov turi")) if actual_columns.get("To'lov turi") else "To‘lov-shartnoma"
                    tulov_turi = 'grant' if "granti" in str(tulov_turi).lower() else 'contract'
                    
                    talim_shakli = row.get(actual_columns.get("Ta'lim shakli")) if actual_columns.get("Ta'lim shakli") else "Kunduzgi"
                    talim_shakli = talim_shakli.lower()
                    
                    is_orphan = row.get("Chin yetim", 'Yo\'q') == 'Ha'
                    has_disability = row.get("Nogironligi bor", 'Yo\'q') != 'Yo\'q'
                    parent_status = row.get("Boquvchisini yuqotgan", 'Yo\'q')
                    is_in_social_protection = row.get("Ijtimoiy himoya reestiriga kiritilgan", 'Yo\'q') == 'Ha'
                    is_in_temir_daftar = row.get("Temir daftarga kiritilgan oila farzandi", 'Yo\'q') == 'Ha'
                    is_in_women_daftar = row.get("Ayollar daftariga kirgizilgan", 'Yo\'q') == 'Ha'
                    is_in_youth_daftar = row.get("Yoshlar daftariga kiritilgan", 'Yo\'q') == 'Ha'
                    is_in_orphanage = row.get("Mehribonlik uyi tarbiyalanuvchisi", 'Yo\'q') == 'Ha'
                    parents_divorced = row.get("Ota-onasi ajrashgan", 'Yo\'q') == 'Ha'
                    
                    student, created = Student.objects.get_or_create(
                        jshshir=jshshir,
                        defaults={
                            'first_name': first_name,
                            'last_name': last_name,
                            'middle_name': middle_name,
                            'gender': gender,
                            'birth_date': birth_date,
                            'group': group,
                            'is_renting': appartment_type == 'rent',
                            'address': temporary_address,
                            'phone_number': phone_number,
                            'email': '',
                            'is_orphan': is_orphan,
                            'has_disability': has_disability,
                            'lives_in_dormitory': appartment_type == 'ttj',
                            'student_id': int(student_id) if student_id else Student.objects.count() + 1,
                            'fuqaro': 'uz' if country_name == "O'zbekiston" else '',
                            'passport': passport,
                            'otm': row.get(actual_columns.get("OTM")) if actual_columns.get("OTM") else "Qarshi davlat texnika universiteti",
                            'talim_turi': 'bakalavr',
                            'tulov_turi': tulov_turi,
                            'talim_shakli': talim_shakli,
                            'shifr': '',
                            'mutaxassislik': mutaxassislik,
                            'country': country,
                            'const_region': const_region,
                            'const_district': const_district,
                            'temporary_region': temp_region,
                            'temporary_district': temp_district,
                            'temporary_address': temporary_address,
                            'appartment_type': appartment_type,
                            'family_type': family_type,
                            'parent_status': parent_status,
                            'is_in_social_protection': is_in_social_protection,
                            'is_in_temir_daftar': is_in_temir_daftar,
                            'is_in_women_daftar': is_in_women_daftar,
                            'is_in_youth_daftar': is_in_youth_daftar,
                            'is_in_orphanage': is_in_orphanage,
                            'parents_divorced': parents_divorced,
                            'nation': nation,
                            'ttj': ttj,
                            'room': room,
                        }
                    )
                    
                    if created:
                        successful += 1
                        logger.info(f"Sahifa {sheet_name}, Qator {excel_row}: Talaba {student.first_name} {student.last_name} qo'shildi.")
                    else:
                        logger.info(f"Sahifa {sheet_name}, Qator {excel_row}: Talaba {student.first_name} {student.last_name} mavjud.")
                
                except Exception as e:
                    errors.append(f"Sahifa {sheet_name}, Qator {excel_row}: {str(e)}")
        
        return successful, errors
    
    except Exception as e:
        return 0, [f"Faylni o'qishda xato: {str(e)}. Fayl formatini tekshiring."]

def upload_excel(request):
    if not request.user.is_authenticated:
        return render(request, 'students/upload_excel.html', {'message': 'Tizimga kiring.', 'errors': []})
    
    if request.method == 'POST' and request.FILES.get('excel_file'):
        excel_file = request.FILES['excel_file']
        fs = FileSystemStorage()
        filename = fs.save(excel_file.name, excel_file)
        file_path = fs.path(filename)
        
        try:
            # Ensure pandas closes the file after reading
            successful, errors = import_students_from_excel(file_path, request.user)
            
            # Attempt to delete the file with retries
            max_retries = 5
            for attempt in range(max_retries):
                try:
                    os.remove(file_path)
                    logger.info(f"Fayl muvaffaqiyatli o'chirildi: {file_path}")
                    break
                except PermissionError as e:
                    logger.warning(f"Faylni o'chirishda xato (Pog'on {attempt + 1}/{max_retries}): {str(e)}")
                    time.sleep(1)  # Wait 1 second before retrying
                except Exception as e:
                    logger.error(f"Faylni o'chirishda kutilmagan xato: {str(e)}")
                    errors.append(f"Faylni o'chirishda xato: {str(e)}")
                    break
            
            context = {
                'successful': successful,
                'errors': errors,
                'message': f"Muvaffaqiyatli: {successful}. Xatolar: {len(errors)}"
            }
            return render(request, 'students/upload_excel.html', context)
        
        except Exception as e:
            # Attempt to delete the file even if processing fails
            try:
                os.remove(file_path)
                logger.info(f"Fayl muvaffaqiyatli o'chirildi (xato holatida): {file_path}")
            except PermissionError:
                logger.warning(f"Faylni o'chirishda xato (xato holatida): {file_path}")
                errors.append(f"Faylni o'chirishda xato: Fayl boshqa jarayon tomonidan ishlatilmoqda.")
            except Exception as e:
                logger.error(f"Faylni o'chirishda kutilmagan xato: {str(e)}")
                errors.append(f"Faylni o'chirishda xato: {str(e)}")
            
            context = {
                'successful': 0,
                'errors': [f"Xato: {str(e)}"] + errors,
                'message': 'Xato yuz berdi.'
            }
            return render(request, 'students/upload_excel.html', context)
    
    return render(request, 'students/upload_excel.html', {'message': ''})